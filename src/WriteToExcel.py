from openpyxl import Workbook
import os
import xlsxwriter as xlsx
from Config import *
import datetime

def getPath(fileName):
    desktop = os.path.normpath(os.path.expanduser("~/Desktop"))
    format = '.xlsx'
    file = desktop + path + fileName + format
    os.makedirs(os.path.dirname(file), exist_ok=True)
    return file

def format_number(number_str):
    number = int(number_str)
    formatted_number = '{:,.0f}'.format(number).replace(",", " ")
    return str(formatted_number)

def getOccurences(df, column_name, value):
    map = {}
    for val in value:
        map[val] = (df[column_name] == val).sum()
    return map

def get_datetime(df, i):
    date = str(df.iloc[i]['Datum'])
    time = str(df.iloc[i]['Tid'])
    year = '20' + date[0:2]
    month = date[2:4]
    day = date[4:6]
    hour = time[0:2]
    minute = time[3:5]
    return datetime.datetime(year=int(year), month=int(month), day=int(day), hour=int(hour), minute=int(minute), second=0)

def numberOfDigits(personnummer):
    count = 0
    for char in personnummer:
        if char.isdigit():
            count += 1
    return count

def half_hour_diff(df, timediff, index):
    dateA = get_datetime(df, index)
    dateB = get_datetime(df, index - 1)
    difference = abs(dateB - dateA)
    correct_timeDiff = (difference  < datetime.timedelta(minutes=timediff+1))
    correct_task = str(df.iloc[index-1]['Tjänst']) == 'Blod' and str(df.iloc[index]['Tjänst']) == 'Blod'
    correct_district = str(df.iloc[index-1]['Distrikt']) == str(df.iloc[index]['Distrikt'])

    return correct_timeDiff and correct_task and correct_district


def write(fileName, df, place):

    if place != 'misnamed':
        df.sort_values(by=['Datum', 'Tid'], inplace=True)

    desktop = os.path.normpath(os.path.expanduser("~/Desktop"))
    file = desktop + path + fileName + file_format
    os.makedirs(os.path.dirname(file), exist_ok=True)

    if not os.path.exists(file):
        create_xls_file(file)

    discount_count = 0
    # Create a new workbook and sheet
    workbook = xlsx.Workbook(file)
    workbook = xlsx.Workbook(file, {'nan_inf_to_errors': True})
    sheet = workbook.add_worksheet('Sheet2')
    yellow_format = workbook.add_format({'bg_color': 'yellow', 'font_size': 14, 'bold':True, 'border': 1})

    column_width = 20  # 20 characters wide
    sheet.set_column(0, len(df.columns) - 1, column_width)
    header_format = workbook.add_format({'font_size': 20})  # Adjust font size as needed
    small_text = workbook.add_format({'font_size': 10})
    light_blue = workbook.add_format({'bg_color': '#ADD8E6', 'border': 1})
    header_light_blue = workbook.add_format({'bg_color': '#ADD8E6', 'border': 1, 'bold':True})
    error_format = workbook.add_format({'bg_color': 'red', 'font_size': 14, 'bold':True, 'border': 1})
    black_borders = workbook.add_format({'border': 1})

    sheet.write(0, 0, str(place), header_format)
    sheet.write(0, 1, 'skapades: ' + str(datetime.date.today()), small_text)

    # Write column names to the first row
    for j, col_name in enumerate(df.columns):
        #if j == 0 or j == len(df.columns) - 1:  # First and last columns    
        sheet.write(startWrite-1, j, col_name, yellow_format)

    map = getOccurences(df, 'Tjänst', [value for value in taskMapping.values()])

    # Write the occurences of the different tasks
    sheet.write(startWrite-3, 0, 'Antal', header_light_blue)
    sheet.write(startWrite-2, 0, 'Totalt Pris', header_light_blue)

    if place != 'misnamed': 
        for index in range(1, len(df)):
            if half_hour_diff(df, 30, index):
                df.loc[index, 'Kostnad'] = rabatt + ' kr'
                discount_count += 1
                print(df)
    
    if place == krim:
        sheet.write(startWrite-4, 1, 'jourläkare', header_light_blue)
        sheet.write(startWrite-3, 1, df.shape[0], header_light_blue)
        sheet.write(startWrite-2, 1, df.shape[0] * int(price_place_task[place]['Jourläkare']), header_light_blue)    
    else:
        for j, task in enumerate(map.keys()):
            if task not in price_place_task[place]:
                continue
            elif task == 'Blod':
                price = price_place_task[place][task]
                sheet.write(startWrite-4, 1 + j, (task +'(Pris='+str(price)+'kr)'), header_light_blue)
                sheet.write(startWrite-3, 1 + j, str(map[task]-discount_count)+',   r: '+str(discount_count), light_blue)
                if str(price).isdigit():
                    sheet.write(startWrite-2, 1 + j, format_number( str((int(price) * (int(map[task])-discount_count)) + (discount_count* int(rabatt))) ), light_blue)
                else:
                    sheet.write(startWrite-2, 1 + j, price, light_blue)
            else:
                price = price_place_task[place][task]
                sheet.write(startWrite-4, 1 + j, (task +'(Pris='+str(price)+'kr)'), header_light_blue)
                sheet.write(startWrite-3, 1 + j, map[task], light_blue)
                if str(price).isdigit():
                    sheet.write(startWrite-2, 1 + j, format_number( str((int(price) * (int(map[task])))) ), light_blue)
                else:
                    sheet.write(startWrite-2, 1 + j, price, light_blue)

    # Write the DataFrame data
    for i, row in enumerate(df.values):
        for j, value in enumerate(row):
            if j == 0 and place != 'misnamed':
                value = convert_date(str(df['Datum'].iloc[i]))
            elif j == 0:
                value = str(df['Datum'].iloc[i])
            if place == 'misnamed' and not valid_cell(df.iloc[i], str(df.columns[j])):
                sheet.write(startWrite + i, j, value, error_format)
            else:
                sheet.write(startWrite + i, j, value)  # Start writing data from the third row

    # Save the workbook to the file
    workbook.close()
def is_valid_time_format(time):
    char = ['.', ':']
    if len(time) == 4 and numberOfDigits(time) == 4:
        if int(time[0]) == 2 and int(time[1]) < 4 and int(time[2]) < 6:
            return True
        elif int(time[0]) < 2 and int(time[1]) <= 9 and int(time[2]) < 6:
            return True        
    elif numberOfDigits(time) == 4 and len(time) == 5 and str(time[2]) in char:
        if int(time[0]) == 2 and int(time[1]) < 4 and int(time[3]) < 6:
            return True
        elif int(time[0]) < 2 and int(time[1]) <= 9 and int(time[3]) < 6:
            return True
    elif numberOfDigits(time) == 3 and len(time) == 4 and str(time[1]) in char:
        if int(time[0]) <= 9 and int(time[2]) < 6 and int(time[3]) <= 9:
            return True
    return False

def valid_cell(row, cell):
    if (isinstance(row, int) or isinstance(row, str)) or str(row['Distrikt']).lower() == "":
        return False
    if cell == 'Distrikt' and not str(row['Distrikt']).lower() in placeMapping:
        return False
    if cell == 'Tid' and not is_valid_time_format(str(row['Tid'])):
        return False
    if cell == 'Tjänst' and str(row['Tjänst']).lower() not in taskMapping:
        return False
    if cell == 'Pers.nr.' and numberOfDigits(str(row['Pers.nr.'])) != 8 and numberOfDigits(str(row['Pers.nr.'])) != 6 and numberOfDigits(str(row['Pers.nr.'])) != 10 and str(row['Pers.nr.']).replace(" ", "").lower() != 'okänd':
        return False
    if cell == 'Datum' and not valid_date(str(row['Datum'])):
        return False
    return True

def valid_date(date_str):
    if (numberOfDigits(date_str) != 6):
        return False
    if (numberOfDigits(date_str) == 8) and date_str[0] == '2' and date_str[1] == '1':
        date_str = date_str[2:]
    if date_str.lower() == '':
        return False
    if date_str[5] == '0' and date_str[4] == '0':
        return False
    if date_str[2] == '0' and str(date_str)[3] == '0':
        return False

    return True

def create_xls_file(file_path):
    # Create an empty file
    open(file_path, 'w').close()

def convert_date(date):
    return str(date[4:6]) + '/' + str(date[2:4]) + '-' + '20' + str(date[0:2])

def delete_contents(file):
    wb = Workbook()
    ws = wb.active
    ws = wb.create_sheet("Sheet1")    
    ws.delete_rows(1, ws.max_row)
    ws.delete_cols(1, ws.max_column)
    
    # Save the changes to the file
    wb.save(file)
